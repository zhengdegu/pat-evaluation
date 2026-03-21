#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Docker 启动时自动导入 生物医药.xlsx 到 Elasticsearch"""

import os
import sys
import time

def wait_for_es(es_url, timeout=120):
    """等待 ES 就绪"""
    import urllib.request
    print("等待 Elasticsearch 启动...")
    start = time.time()
    while time.time() - start < timeout:
        try:
            req = urllib.request.urlopen(es_url + '/_cluster/health', timeout=5)
            if req.status == 200:
                print("Elasticsearch 已就绪")
                return True
        except Exception:
            pass
        time.sleep(2)
    print("等待 Elasticsearch 超时")
    return False


def main():
    es_url = os.environ.get('ESURL', 'http://127.0.0.1:9200')
    xlsx_path = os.environ.get('IMPORT_XLSX', '/app/backend/生物医药.xlsx')
    index_name = 'patent_new2'
    doc_type = 'content'

    if not os.path.exists(xlsx_path):
        print(f"文件不存在: {xlsx_path}，跳过自动导入")
        return

    if not wait_for_es(es_url):
        sys.exit(1)

    # 检查索引中是否已有数据，避免重复导入
    from elasticsearch import Elasticsearch
    es = Elasticsearch(es_url)

    try:
        count = es.count(index=index_name)['count']
        if count > 0:
            print(f"索引 {index_name} 已有 {count} 条数据，跳过导入")
            return
    except Exception:
        pass  # 索引可能还不存在，继续

    # 等待索引创建（init-es.sh 会创建）
    for _ in range(30):
        if es.indices.exists(index=index_name):
            break
        time.sleep(2)

    print(f"开始导入 {xlsx_path} ...")
    import openpyxl
    from elasticsearch.helpers import bulk

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        print("Excel 文件为空")
        wb.close()
        return

    headers = [str(h).strip() if h else '' for h in rows[0]]
    print(f"表头: {headers}")
    print(f"数据行数: {len(rows) - 1}")

    ES_TRADE_FIELD = '转化收益（万元）'
    batch = []
    success_total = 0

    for row in rows[1:]:
        if not any(row):
            continue
        row_dict = dict(zip(headers, row))
        doc = {}
        for k, v in row_dict.items():
            if v is None:
                v = ''
            k = str(k).strip()
            if k == 'IPC分类号':
                s = str(v).strip() if v else ''
                for sep in [';', '；', ',', '，', '\n']:
                    if sep in s:
                        doc[k] = [x.strip() for x in s.split(sep) if x.strip()]
                        break
                else:
                    doc[k] = [s] if s else []
            elif '转化收益' in k:
                try:
                    doc[ES_TRADE_FIELD] = float(v) if v != '' else 0.0
                except (ValueError, TypeError):
                    doc[ES_TRADE_FIELD] = 0.0
            elif k == '专利号':
                doc['专利号'] = str(v).strip()
                if '申请号' not in doc:
                    doc['申请号'] = str(v).strip()
            elif k == '申请号':
                doc['申请号'] = str(v).strip()
            else:
                doc[k] = str(v).strip() if v else ''

        if not doc.get('专利号') and not doc.get('申请号'):
            continue

        batch.append({'_index': index_name, '_type': doc_type, '_source': doc})

        if len(batch) >= 200:
            ok, failed = bulk(es, batch, raise_on_error=False)
            success_total += ok
            batch = []

    if batch:
        ok, failed = bulk(es, batch, raise_on_error=False)
        success_total += ok

    wb.close()
    print(f"导入完成，共 {success_total} 条记录")


if __name__ == '__main__':
    main()
