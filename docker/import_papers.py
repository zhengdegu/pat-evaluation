#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Docker 启动时自动导入论文数据到 Elasticsearch"""

import os
import sys
import time


def wait_for_es(es_url, timeout=120):
    """等待 ES 就绪"""
    import urllib.request
    print("等待 Elasticsearch 启动...")
    start = time.time()
    while time.time() - start < timeout:
        try:
            req = urllib.request.urlopen(es_url + '/_cluster/health', timeout=5)
            if req.status == 200:
                print("Elasticsearch 已就绪")
                return True
        except Exception:
            pass
        time.sleep(2)
    print("等待 Elasticsearch 超时")
    return False


def main():
    es_url = os.environ.get('ESURL', 'http://127.0.0.1:9200')
    index_name = 'paper_data'
    doc_type = 'content'

    # 论文 Excel 文件路径
    paper_files = [
        os.environ.get('PAPER_EN_XLSX', '/app/backend/data/论文数据_英文.xlsx'),
        os.environ.get('PAPER_CN_XLSX', '/app/backend/data/论文数据_中文.xlsx'),
    ]

    # 检查是否有文件存在
    existing_files = [f for f in paper_files if os.path.exists(f)]
    if not existing_files:
        print("未找到论文数据文件，跳过导入")
        return

    if not wait_for_es(es_url):
        sys.exit(1)

    from elasticsearch import Elasticsearch
    from elasticsearch.helpers import bulk
    import openpyxl

    es = Elasticsearch(es_url)

    # 检查索引中是否已有数据
    try:
        count = es.count(index=index_name)['count']
        if count > 0:
            print(f"索引 {index_name} 已有 {count} 条数据，跳过导入")
            return
    except Exception:
        pass

    # 等待索引创建（init-es.sh 会创建）
    for _ in range(30):
        if es.indices.exists(index=index_name):
            break
        time.sleep(2)

    total_imported = 0

    for xlsx_path in existing_files:
        print(f"导入 {xlsx_path} ...")

        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            print(f"  文件为空，跳过")
            wb.close()
            continue

        headers = [str(h).strip() if h else '' for h in rows[0]]
        print(f"  表头: {headers}")
        print(f"  数据行数: {len(rows) - 1}")

        batch = []
        file_count = 0

        for row in rows[1:]:
            if not any(row):
                continue

            doc = {}
            for k, v in zip(headers, row):
                k = str(k).strip()
                if not k:
                    continue
                if v is None:
                    v = ''

                if k == '被引用次数':
                    try:
                        doc[k] = int(float(v)) if v else 0
                    except (ValueError, TypeError):
                        doc[k] = 0
                else:
                    doc[k] = str(v).strip()

            if not doc.get('论文名称'):
                continue

            batch.append({
                '_index': index_name,
                '_type': doc_type,
                '_source': doc,
            })

            if len(batch) >= 500:
                ok, failed = bulk(es, batch, raise_on_error=False)
                file_count += ok
                batch = []

        if batch:
            ok, failed = bulk(es, batch, raise_on_error=False)
            file_count += ok

        wb.close()
        total_imported += file_count
        print(f"  导入 {file_count} 条")

    # 刷新索引
    try:
        es.indices.refresh(index=index_name)
    except Exception:
        pass

    print(f"论文导入完成，共 {total_imported} 条")


if __name__ == '__main__':
    main()
