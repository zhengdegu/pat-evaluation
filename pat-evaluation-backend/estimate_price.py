#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
基于专利特征的转化收益估算模型

估算依据（生物医药领域）：
1. 专利类型：发明 vs 实用新型
2. IPC分类号：不同技术领域的市场价值差异
3. 法律状态：有效 > 审中 > 失效
4. 申请人类型：企业 vs 高校/研究机构 vs 其他
5. 申请年份 → 剩余有效期
6. 是否有优先权（国际布局）
7. 引用专利数量（技术关联度）
8. 引用文献数量（研究深度）
"""

import openpyxl
import datetime
import random
import math

random.seed(42)  # 可复现

# ============================================================
# 1. IPC 分类号对应的基础价格区间（万元）
#    参考：中国技术交易所、各省技术产权交易中心公开统计
# ============================================================
IPC_BASE_PRICE = {
    # 药物制剂/医药 — 高价值
    'A61K': (80, 350),
    'A61P': (80, 320),
    # 基因工程/微生物 — 高价值
    'C12N': (100, 400),
    'C12Q': (60, 250),
    'C12P': (50, 200),
    'C12M': (40, 150),
    # 检测/分析 — 中高价值
    'G01N': (50, 200),
    # 有机化学 — 中高价值
    'C07H': (60, 280),
    'C07D': (60, 260),
    'C07K': (70, 300),  # 蛋白质/多肽
    'C07C': (40, 180),
    # 食品 — 中等价值
    'A23L': (30, 120),
    'A23K': (20, 80),
    # 农业/种植 — 中等价值
    'A01G': (20, 100),
    'A01K': (20, 90),
    # 高分子/材料
    'C08B': (30, 130),
}
DEFAULT_BASE_PRICE = (30, 150)

# ============================================================
# 2. 各因素的修正系数
# ============================================================

def get_ipc_base(ipc_str):
    """根据IPC分类号取基础价格区间"""
    if not ipc_str:
        return DEFAULT_BASE_PRICE
    prefix = str(ipc_str).strip()[:4]
    return IPC_BASE_PRICE.get(prefix, DEFAULT_BASE_PRICE)


def patent_type_factor(ptype):
    """专利类型系数：发明专利价值远高于实用新型"""
    if ptype == '发明':
        return 1.0
    elif ptype == '实用新型':
        return 0.35
    else:
        return 0.2


def legal_status_factor(status):
    """法律状态系数"""
    if status == '有效':
        return 1.0
    elif status == '审中':
        return 0.6  # 未授权，有不确定性
    elif status == '失效':
        return 0.05  # 几乎无价值，但技术方案本身可能有参考价值
    return 0.5


def applicant_type_factor(applicant):
    """申请人类型系数"""
    if not applicant:
        return 0.6
    applicant = str(applicant)
    # 大型药企/生物科技公司 — 转化能力强
    big_companies = ['华大', '药明', '恒瑞', '百济', '信达', '君实', '科兴',
                     '国药', '中国医药', '复星', '石药', '齐鲁', '正大天晴',
                     '康泰', '智飞', '沃森', '康希诺', '艾博']
    for bc in big_companies:
        if bc in applicant:
            return 1.3

    if '公司' in applicant or '企业' in applicant or '集团' in applicant:
        return 1.1
    elif '大学' in applicant or '学院' in applicant:
        return 0.85  # 高校专利转化率偏低
    elif '研究' in applicant or '科学院' in applicant:
        return 0.9
    elif '医院' in applicant:
        return 0.8
    return 0.7


def remaining_life_factor(apply_date_str, patent_type):
    """剩余有效期系数"""
    if not apply_date_str:
        return 0.5
    try:
        date_str = str(apply_date_str).strip()
        # 尝试多种日期格式
        for fmt in ['%Y-%m-%d', '%Y%m%d', '%Y/%m/%d', '%Y.%m.%d',
                     '%Y-%m-%d %H:%M:%S', '%Y-%m-%d 00:00:00']:
            try:
                apply_date = datetime.datetime.strptime(date_str[:10], fmt[:len(date_str[:10])+2])
                break
            except ValueError:
                continue
        else:
            # 尝试只取前10个字符
            try:
                apply_date = datetime.datetime.strptime(date_str[:10], '%Y-%m-%d')
            except ValueError:
                return 0.5

        max_years = 20 if patent_type == '发明' else 10
        elapsed = (datetime.datetime.now() - apply_date).days / 365.25
        remaining = max(0, max_years - elapsed)
        ratio = remaining / max_years

        if remaining <= 0:
            return 0.05
        elif ratio > 0.8:
            return 1.1  # 新专利，大部分有效期剩余
        elif ratio > 0.5:
            return 1.0
        elif ratio > 0.3:
            return 0.7
        else:
            return 0.4
    except Exception:
        return 0.5


def priority_factor(priority_str):
    """优先权系数：有国际优先权说明有海外布局，价值更高"""
    if not priority_str or str(priority_str).strip() in ('', 'None', 'nan'):
        return 1.0
    return 1.25


def citation_factor(cite_patents, cite_refs):
    """引用因素系数：引用越多说明技术关联度和研究深度越高"""
    cite_count = 0
    if cite_patents and str(cite_patents).strip() not in ('', 'None', 'nan'):
        cite_count += len([x for x in str(cite_patents).replace('；', ';').split(';') if x.strip()])

    ref_count = 0
    if cite_refs and str(cite_refs).strip() not in ('', 'None', 'nan'):
        ref_count += len([x for x in str(cite_refs).replace('；', ';').split(';') if x.strip()])

    total = cite_count + ref_count
    if total == 0:
        return 0.85
    elif total <= 3:
        return 1.0
    elif total <= 8:
        return 1.1
    else:
        return 1.2


def estimate_patent_price(row):
    """
    综合估算单条专利的转化收益（万元）

    row: Excel行数据 tuple
    列索引:
      0=专利号, 1=专利名, 2=主分类号, 3=IPC分类号, 4=申请人,
      5=当前权利人, 6=发明人, 7=公开日, 8=公开号, 9=代理机构,
      10=代理人, 11=申请日, 12=申请人地址, 13=优先权, 14=国省代码,
      15=摘要, 16=主权项, 17=引用专利, 18=引用文献, 19=法律状态,
      20=专利类型, 21=转化收益(万元)
    """
    ipc = row[3]
    applicant = row[4]
    apply_date = row[11]
    priority = row[13]
    cite_patents = row[17]
    cite_refs = row[18]
    legal_status = str(row[19]).strip() if row[19] else '未知'
    patent_type = str(row[20]).strip() if row[20] else '发明'

    # 1. 基础价格区间
    low, high = get_ipc_base(ipc)
    base_price = (low + high) / 2

    # 2. 各因素修正
    f1 = patent_type_factor(patent_type)
    f2 = legal_status_factor(legal_status)
    f3 = applicant_type_factor(applicant)
    f4 = remaining_life_factor(apply_date, patent_type)
    f5 = priority_factor(priority)
    f6 = citation_factor(cite_patents, cite_refs)

    # 综合修正
    combined = f1 * f2 * f3 * f4 * f5 * f6
    price = base_price * combined

    # 加入随机波动 ±15%，模拟市场不确定性
    noise = random.uniform(0.85, 1.15)
    price = price * noise

    # 失效专利给极低值
    if legal_status == '失效':
        price = max(price, 0.5)
        price = min(price, 10)

    # 四舍五入到整数
    price = round(max(price, 1), 0)

    return price


def main():
    xlsx_path = '生物医药.xlsx'
    print(f"读取 {xlsx_path} ...")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = rows[0]
    trade_col_idx = None
    for i, h in enumerate(headers):
        if h and '转化收益' in str(h):
            trade_col_idx = i
            break

    if trade_col_idx is None:
        print("找不到转化收益列！")
        return

    print(f"转化收益列索引: {trade_col_idx} (列名: {headers[trade_col_idx]})")

    # 重新打开用于写入（非 read_only 模式）
    wb_write = openpyxl.load_workbook(xlsx_path)
    ws_write = wb_write[wb_write.sheetnames[0]]

    # 统计
    total = 0
    status_counts = {'有效': 0, '审中': 0, '失效': 0, '其他': 0}
    price_sum = 0

    for row_idx, row in enumerate(rows[1:], start=2):  # Excel行号从2开始（跳过表头）
        price = estimate_patent_price(row)
        # 写入 Excel（列号从1开始）
        ws_write.cell(row=row_idx, column=trade_col_idx + 1, value=price)

        total += 1
        price_sum += price
        ls = str(row[19]).strip() if row[19] else '其他'
        if ls in status_counts:
            status_counts[ls] += 1
        else:
            status_counts['其他'] += 1

    wb_write.save(xlsx_path)
    print(f"\n完成！共更新 {total} 条专利的转化收益")
    print(f"平均估算价格: {price_sum/total:.1f} 万元")
    print(f"法律状态分布: {status_counts}")

    # 输出前20条看看效果
    print("\n=== 前20条估算结果 ===")
    wb_check = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws_check = wb_check[wb_check.sheetnames[0]]
    check_rows = list(ws_check.iter_rows(min_row=1, values_only=True))
    for i, row in enumerate(check_rows[1:21], 1):
        patid = row[0]
        name = str(row[1])[:35] if row[1] else ''
        ptype = row[20]
        status = row[19]
        price = row[trade_col_idx]
        print(f"  {i:2d}. {patid} | {name:35s} | {ptype} | {status} | {price} 万元")
    wb_check.close()


if __name__ == '__main__':
    main()
