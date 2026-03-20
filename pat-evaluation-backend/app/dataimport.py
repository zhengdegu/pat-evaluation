# -*- coding=utf-8 -*-
"""数据导入模块 — 支持上传 Excel 文件并导入到 Elasticsearch"""

import os
import tempfile
from flask import Blueprint, jsonify, request, current_app as app
from .utils import get_esclient, get_index_name

import_bp = Blueprint('dataimport', __name__, url_prefix='/api/import')

# Excel 列名到 ES 字段的映射（保持一致）
FIELD_ORDER = [
    '专利号', '专利名', '主分类号', 'IPC分类号', '申请人', '当前权利人',
    '发明人', '公开日', '公开号', '代理机构', '代理人', '申请日',
    '申请人地址', '优先权', '国省代码', '摘要', '主权项',
    '引用专利', '引用文献', '法律状态', '专利类型', '转化收益(万元)'
]

# ES 中转化收益字段名（注意括号是中文全角）
ES_TRADE_FIELD = '转化收益（万元）'


def _process_ipc(val):
    """将 IPC 分类号转为列表"""
    if not val:
        return []
    s = str(val).strip()
    for sep in [';', '；', ',', '，', '\n']:
        if sep in s:
            return [x.strip() for x in s.split(sep) if x.strip()]
    return [s]


def _row_to_doc(row_dict):
    """将一行数据转为 ES 文档"""
    doc = {}
    for k, v in row_dict.items():
        if v is None:
            v = ''
        k_stripped = str(k).strip()

        if k_stripped == 'IPC分类号':
            doc['IPC分类号'] = _process_ipc(v)
        elif k_stripped == '转化收益(万元)' or k_stripped == '转化收益（万元）':
            try:
                doc[ES_TRADE_FIELD] = float(v) if v != '' else 0.0
            except (ValueError, TypeError):
                doc[ES_TRADE_FIELD] = 0.0
        elif k_stripped == '申请号':
            # 有些 Excel 里列名叫"专利号"但实际是申请号
            doc['申请号'] = str(v).strip()
        elif k_stripped == '专利号':
            doc['专利号'] = str(v).strip()
            # 同时写入申请号（兼容查询）
            if '申请号' not in doc:
                doc['申请号'] = str(v).strip()
        else:
            doc[k_stripped] = str(v).strip() if v else ''
    return doc


@import_bp.route('/upload', methods=['POST'])
def upload_excel():
    """上传 Excel 文件并导入到 ES"""
    if 'file' not in request.files:
        return jsonify(error='未选择文件', error_code='no_file'), 400

    f = request.files['file']
    if not f.filename:
        return jsonify(error='未选择文件', error_code='no_file'), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return jsonify(error='仅支持 .xlsx / .xls 格式', error_code='bad_format'), 400

    # 保存到临时文件
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
    try:
        f.save(tmp.name)
        tmp.close()

        import openpyxl
        wb = openpyxl.load_workbook(tmp.name, read_only=True)
        ws = wb[wb.sheetnames[0]]

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return jsonify(error='文件为空或仅有表头', error_code='empty'), 400

        headers = [str(h).strip() if h else '' for h in rows[0]]
        app.logger.info("Excel headers: %s", headers)

        es = get_esclient()
        index_name = get_index_name('patent')
        doc_type = 'content'

        success_count = 0
        fail_count = 0
        errors = []

        from elasticsearch.helpers import bulk

        batch = []
        for idx, row in enumerate(rows[1:], start=2):
            row_dict = dict(zip(headers, row))
            # 跳过完全空行
            if not any(row):
                continue
            doc = _row_to_doc(row_dict)
            if not doc.get('专利号') and not doc.get('申请号'):
                fail_count += 1
                errors.append(f"第{idx}行: 缺少专利号/申请号")
                continue

            batch.append({
                '_index': index_name,
                '_type': doc_type,
                '_source': doc
            })

            if len(batch) >= 200:
                ok, failed = bulk(es, batch, raise_on_error=False)
                success_count += ok
                if failed:
                    fail_count += len(failed)
                batch = []

        # 剩余
        if batch:
            ok, failed = bulk(es, batch, raise_on_error=False)
            success_count += ok
            if failed:
                fail_count += len(failed)

        wb.close()
        return jsonify(
            error_code='success',
            success_count=success_count,
            fail_count=fail_count,
            total=len(rows) - 1,
            errors=errors[:20]  # 最多返回前20条错误
        )

    except Exception as e:
        app.logger.error("Import error: %s", e)
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify(error=str(e), error_code='internal_error'), 500
    finally:
        os.unlink(tmp.name)


@import_bp.route('/status')
def import_status():
    """查询当前 ES 中专利数据总量"""
    try:
        es = get_esclient()
        index_name = get_index_name('patent')
        count = es.count(index=index_name)['count']
        return jsonify(error_code='success', count=count)
    except Exception as e:
        app.logger.error("Status check error: %s", e)
        return jsonify(error_code='error', error=str(e), count=0)
