#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将论文 Excel 数据导入 Elasticsearch paper_data 索引

用法：
    python scripts/import_papers_to_es.py [--es-url http://10.0.9.80:9200]
"""

import os
import sys
import argparse
import logging

import openpyxl
from elasticsearch import Elasticsearch
from elasticsearch.helpers import bulk

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INDEX_NAME = 'paper_data'
DOC_TYPE = 'content'

# 两个论文 Excel 文件
PAPER_FILES = [
    os.path.join(BASE_DIR, 'data', '论文数据_英文.xlsx'),
    os.path.join(BASE_DIR, 'data', '论文数据_中文.xlsx'),
]


def ensure_index(es: Elasticsearch):
    """确保 paper_data 索引存在，不存在则创建"""
    if es.indices.exists(index=INDEX_NAME):
        logger.info(f'索引 {INDEX_NAME} 已存在')
        count = es.count(index=INDEX_NAME)['count']
        logger.info(f'当前文档数: {count}')
        return

    mapping = {
        'mappings': {
            DOC_TYPE: {
                'properties': {
                    '论文名称': {'type': 'text', 'analyzer': 'ik_max_word'},
                    '作者': {'type': 'text', 'analyzer': 'ik_max_word'},
                    '摘要': {'type': 'text', 'analyzer': 'ik_max_word'},
                    '发表年份': {'type': 'keyword'},
                    '被引用次数': {'type': 'integer'},
                    '期刊': {'type': 'text', 'analyzer': 'ik_max_word'},
                    'DOI': {'type': 'keyword'},
                    'PMID': {'type': 'keyword'},
                    '数据来源': {'type': 'keyword'},
                    '关联专利号': {'type': 'keyword'},
                    '关联发明人': {'type': 'text', 'analyzer': 'ik_max_word'},
                    '关联申请人': {'type': 'text', 'analyzer': 'ik_max_word'},
                }
            }
        }
    }

    # 如果没有 ik 分词器，退回 standard
    try:
        es.indices.create(index=INDEX_NAME, body=mapping)
        logger.info(f'创建索引 {INDEX_NAME}（ik_max_word 分词）')
    except Exception as e:
        if 'ik_max_word' in str(e):
            # 没装 ik 插件，用 standard
            for field in mapping['mappings'][DOC_TYPE]['properties'].values():
                if field.get('analyzer') == 'ik_max_word':
                    field['analyzer'] = 'standard'
            es.indices.create(index=INDEX_NAME, body=mapping)
            logger.info(f'创建索引 {INDEX_NAME}（standard 分词，未安装 ik 插件）')
        else:
            raise


def load_papers_from_excel(xlsx_path: str) -> list:
    """从 Excel 加载论文数据"""
    if not os.path.exists(xlsx_path):
        logger.warning(f'文件不存在: {xlsx_path}')
        return []

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        wb.close()
        return []

    headers = [str(h).strip() if h else '' for h in rows[0]]
    papers = []

    for row in rows[1:]:
        if not any(row):
            continue
        d = dict(zip(headers, row))

        doc = {}
        for k, v in d.items():
            k = str(k).strip()
            if not k:
                continue
            if v is None:
                v = ''

            if k == '被引用次数':
                try:
                    doc[k] = int(float(v)) if v else 0
                except (ValueError, TypeError):
                    doc[k] = 0
            else:
                doc[k] = str(v).strip()

        # 跳过无标题的
        if not doc.get('论文名称'):
            continue

        papers.append(doc)

    wb.close()
    logger.info(f'从 {os.path.basename(xlsx_path)} 加载 {len(papers)} 篇论文')
    return papers


def import_to_es(es: Elasticsearch, papers: list) -> int:
    """批量导入论文到 ES"""
    if not papers:
        return 0

    batch = []
    success_total = 0

    for paper in papers:
        batch.append({
            '_index': INDEX_NAME,
            '_type': DOC_TYPE,
            '_source': paper,
        })

        if len(batch) >= 500:
            ok, failed = bulk(es, batch, raise_on_error=False)
            success_total += ok
            if failed:
                logger.warning(f'{len(failed)} 条导入失败')
            batch = []

    if batch:
        ok, failed = bulk(es, batch, raise_on_error=False)
        success_total += ok
        if failed:
            logger.warning(f'{len(failed)} 条导入失败')

    return success_total


def main():
    parser = argparse.ArgumentParser(description='导入论文数据到 Elasticsearch')
    parser.add_argument('--es-url', default=os.environ.get('ESURL', 'http://10.0.9.80:9200'),
                        help='ES 地址')
    parser.add_argument('--clear', action='store_true', help='导入前清空索引')
    args = parser.parse_args()

    logger.info(f'=== 论文数据导入 ES ===')
    logger.info(f'ES: {args.es_url}')
    logger.info(f'索引: {INDEX_NAME}')

    es = Elasticsearch(args.es_url)

    # 检查 ES 连接
    try:
        info = es.info()
        logger.info(f'ES 版本: {info["version"]["number"]}')
    except Exception as e:
        logger.error(f'无法连接 ES: {e}')
        sys.exit(1)

    # 清空索引（如果指定）
    if args.clear and es.indices.exists(index=INDEX_NAME):
        es.indices.delete(index=INDEX_NAME)
        logger.info(f'已删除索引 {INDEX_NAME}')

    ensure_index(es)

    # 加载并导入
    total_imported = 0
    for xlsx_path in PAPER_FILES:
        papers = load_papers_from_excel(xlsx_path)
        if papers:
            count = import_to_es(es, papers)
            total_imported += count
            logger.info(f'  导入 {count} 篇')

    # 刷新索引
    es.indices.refresh(index=INDEX_NAME)
    final_count = es.count(index=INDEX_NAME)['count']

    logger.info(f'=== 完成 ===')
    logger.info(f'本次导入: {total_imported} 篇')
    logger.info(f'索引总量: {final_count} 篇')


if __name__ == '__main__':
    main()
