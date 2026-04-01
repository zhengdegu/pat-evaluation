#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Phase 1: 检索国外英文论文（CrossRef + Semantic Scholar）
从生物医药.xlsx 提取发明人+申请人，结果存入 论文数据_英文.xlsx

用法：
    python scripts/fetch_papers_en.py [--limit N]
"""

import os
import sys
import time
import re
import logging
import hashlib
import argparse
from typing import List, Dict, Set

import requests
import openpyxl
from openpyxl.styles import Font

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

INPUT_XLSX = os.path.join(BASE_DIR, '生物医药.xlsx')
OUTPUT_XLSX = os.path.join(BASE_DIR, 'data', '论文数据_英文.xlsx')
MAX_PAPERS_PER_PATENT = 10
REQUEST_DELAY = 1.5
MAX_INVENTORS_PER_PATENT = 3

HEADERS_API = {
    'User-Agent': 'PatEvaluation/1.0 (mailto:research@example.com)',
    'Accept': 'application/json',
}

CROSSREF_API = 'https://api.crossref.org/works'
S2_API = 'https://api.semanticscholar.org/graph/v1/paper/search'


def load_patents(xlsx_path: str) -> List[Dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else '' for h in rows[0]]

    patents = []
    for row in rows[1:]:
        d = dict(zip(headers, row))
        patid = str(d.get('专利号', '') or '').strip()
        patent_name = str(d.get('专利名', '') or '').strip()
        inv_str = str(d.get('发明人', '') or '').replace('；', ';')
        app_str = str(d.get('申请人', '') or '').replace('；', ';')
        ipc = str(d.get('IPC分类号', '') or '').strip()

        inventors = [n.strip() for n in inv_str.split(';') if n.strip() and len(n.strip()) >= 2]
        applicant = app_str.split(';')[0].strip() if app_str else ''

        if not inventors:
            continue

        patents.append({
            'patid': patid,
            'patent_name': patent_name,
            'inventors': inventors[:MAX_INVENTORS_PER_PATENT],
            'applicant': applicant,
            'ipc': ipc,
        })

    wb.close()
    logger.info(f'加载 {len(patents)} 条专利')
    return patents


def search_crossref(query: str, limit: int = 5) -> List[Dict]:
    try:
        params = {
            'query': query,
            'rows': limit,
            'select': 'title,author,abstract,published-print,is-referenced-by-count,container-title,DOI',
        }
        resp = requests.get(CROSSREF_API, params=params, headers=HEADERS_API, timeout=20)
        if resp.status_code != 200:
            logger.debug(f'CrossRef {resp.status_code}')
            return []

        papers = []
        for item in resp.json().get('message', {}).get('items', []):
            title = item.get('title', [''])[0] if item.get('title') else ''
            authors = ';'.join(
                f"{a.get('given', '')} {a.get('family', '')}".strip()
                for a in item.get('author', [])
            )
            abstract_raw = item.get('abstract', '')
            abstract = re.sub(r'<[^>]+>', '', abstract_raw)[:500] if abstract_raw else ''
            pub_date = item.get('published-print', {}).get('date-parts', [[None]])[0]
            year = pub_date[0] if pub_date and pub_date[0] else ''
            journal = item.get('container-title', [''])[0] if item.get('container-title') else ''

            if title:
                papers.append({
                    '论文名称': title,
                    '作者': authors,
                    '摘要': abstract,
                    '发表年份': str(year),
                    '被引用次数': item.get('is-referenced-by-count', 0),
                    '期刊': journal,
                    'DOI': item.get('DOI', ''),
                    '数据来源': 'CrossRef',
                })
        return papers
    except Exception as e:
        logger.debug(f'CrossRef error: {e}')
        return []


def search_s2(query: str, limit: int = 5) -> List[Dict]:
    try:
        params = {
            'query': query,
            'limit': limit,
            'fields': 'title,authors,abstract,year,citationCount,journal,externalIds',
        }
        resp = requests.get(S2_API, params=params, headers=HEADERS_API, timeout=20)
        if resp.status_code == 429:
            logger.warning('S2 rate limited, sleeping 60s...')
            time.sleep(60)
            resp = requests.get(S2_API, params=params, headers=HEADERS_API, timeout=20)
        if resp.status_code != 200:
            return []

        papers = []
        for item in resp.json().get('data', []):
            authors = ';'.join(a.get('name', '') for a in item.get('authors', []))
            doi = (item.get('externalIds') or {}).get('DOI', '')
            journal = (item.get('journal') or {}).get('name', '')
            title = item.get('title', '')
            if title:
                papers.append({
                    '论文名称': title,
                    '作者': authors,
                    '摘要': (item.get('abstract') or '')[:500],
                    '发表年份': str(item.get('year', '')),
                    '被引用次数': item.get('citationCount', 0),
                    '期刊': journal,
                    'DOI': doi,
                    '数据来源': 'Semantic Scholar',
                })
        return papers
    except Exception as e:
        logger.debug(f'S2 error: {e}')
        return []


def build_en_queries(patent: Dict) -> List[str]:
    """构建英文检索查询 — 用专利名关键词 + IPC 领域"""
    queries = []
    name = patent['patent_name']

    # 策略1: 直接用专利名（很多生物医药专利名含英文术语）
    if name:
        queries.append(name[:80])

    # 策略2: 申请人英文名（如果有）+ 关键词
    applicant = patent['applicant']
    if applicant:
        queries.append(f'{applicant} {name[:30]}')

    return queries


def paper_key(paper: Dict) -> str:
    title = paper.get('论文名称', '').strip().lower()
    return hashlib.md5(title.encode('utf-8')).hexdigest()


def fetch_papers(patents: List[Dict]) -> List[Dict]:
    all_papers = []
    seen: Set[str] = set()
    total = len(patents)
    s2_ok = True

    for idx, patent in enumerate(patents):
        logger.info(f'[{idx+1}/{total}] {patent["patid"]}')

        queries = build_en_queries(patent)
        found = []

        for q in queries:
            if len(found) >= MAX_PAPERS_PER_PATENT:
                break

            # CrossRef 优先（稳定、不限流）
            results = search_crossref(q, limit=5)
            time.sleep(REQUEST_DELAY)

            # S2 补充
            if len(results) < 3 and s2_ok:
                s2_res = search_s2(q, limit=5)
                if not s2_res and len(results) == 0:
                    # 可能被限流
                    s2_ok = False
                results.extend(s2_res)
                time.sleep(REQUEST_DELAY)

            for p in results:
                k = paper_key(p)
                if k not in seen and p.get('论文名称'):
                    seen.add(k)
                    p['关联专利号'] = patent['patid']
                    p['关联发明人'] = ';'.join(patent['inventors'])
                    p['关联申请人'] = patent['applicant']
                    found.append(p)

        all_papers.extend(found[:MAX_PAPERS_PER_PATENT])

        if found:
            logger.info(f'  -> {len(found)} papers (total {len(all_papers)})')

        # 中间保存
        if (idx + 1) % 50 == 0:
            save_excel(all_papers, OUTPUT_XLSX)
            logger.info(f'  [checkpoint] {len(all_papers)} papers saved')

    return all_papers


def save_excel(papers: List[Dict], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '英文论文'

    cols = ['论文名称', '作者', '摘要', '发表年份', '被引用次数',
            '期刊', 'DOI', '数据来源', '关联专利号', '关联发明人', '关联申请人']

    for ci, cn in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=cn)
        cell.font = Font(bold=True)

    for ri, paper in enumerate(papers, 2):
        for ci, cn in enumerate(cols, 1):
            ws.cell(row=ri, column=ci, value=paper.get(cn, ''))

    widths = [50, 30, 60, 10, 12, 30, 30, 15, 20, 20, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description='英文论文检索（Phase 1）')
    parser.add_argument('--limit', type=int, default=0, help='处理前N条专利（0=全部）')
    args = parser.parse_args()

    logger.info('=== Phase 1: 英文论文检索 ===')

    if not os.path.exists(INPUT_XLSX):
        logger.error(f'找不到: {INPUT_XLSX}')
        sys.exit(1)

    patents = load_patents(INPUT_XLSX)
    if args.limit > 0:
        patents = patents[:args.limit]

    logger.info(f'检索 {len(patents)} 条专利...')
    logger.info(f'预计耗时: ~{len(patents) * 2 * REQUEST_DELAY / 60:.0f} 分钟')

    papers = fetch_papers(patents)
    save_excel(papers, OUTPUT_XLSX)

    logger.info(f'=== 完成: {len(papers)} 篇英文论文 -> {OUTPUT_XLSX} ===')


if __name__ == '__main__':
    main()
