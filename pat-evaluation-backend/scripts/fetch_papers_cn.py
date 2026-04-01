#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Phase 2: 检索中文相关论文（PubMed + CrossRef 补充）
用发明人姓名 + 申请人单位 + 专利关键词检索

PubMed E-utilities 免费，无需 API key，限制 3 req/s（无 key）或 10 req/s（有 key）

用法：
    python scripts/fetch_papers_cn.py [--limit N]
"""

import os
import sys
import time
import re
import logging
import hashlib
import argparse
from typing import List, Dict, Set

import requests
import openpyxl
from openpyxl.styles import Font

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

INPUT_XLSX = os.path.join(BASE_DIR, '生物医药.xlsx')
OUTPUT_XLSX = os.path.join(BASE_DIR, 'data', '论文数据_中文.xlsx')
MAX_PAPERS_PER_PATENT = 8
REQUEST_DELAY = 0.5  # PubMed 允许 3 req/s without key
MAX_INVENTORS_PER_PATENT = 3

HEADERS = {
    'User-Agent': 'PatEvaluation/1.0 (mailto:research@example.com)',
    'Accept': 'application/json',
}

PUBMED_SEARCH = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi'
PUBMED_SUMMARY = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi'
PUBMED_FETCH = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi'

# 常见中文姓氏拼音映射（用于构建 PubMed 查询）
# PubMed 作者格式: "LastName FirstInitial" 如 "Wang J" "Zhang L"
# 中文名: 姓在前，名在后，PubMed 中通常是 "Zhang San" 或 "Zhang S"


def load_patents(xlsx_path: str) -> List[Dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else '' for h in rows[0]]

    patents = []
    for row in rows[1:]:
        d = dict(zip(headers, row))
        patid = str(d.get('专利号', '') or '').strip()
        patent_name = str(d.get('专利名', '') or '').strip()
        inv_str = str(d.get('发明人', '') or '').replace('；', ';')
        app_str = str(d.get('申请人', '') or '').replace('；', ';')
        ipc = str(d.get('IPC分类号', '') or '').strip()
        abstract = str(d.get('摘要', '') or '').strip()

        inventors = [n.strip() for n in inv_str.split(';') if n.strip() and len(n.strip()) >= 2]
        applicant = app_str.split(';')[0].strip() if app_str else ''

        if not inventors:
            continue

        patents.append({
            'patid': patid,
            'patent_name': patent_name,
            'inventors': inventors[:MAX_INVENTORS_PER_PATENT],
            'applicant': applicant,
            'ipc': ipc,
            'abstract': abstract[:200],
        })

    wb.close()
    logger.info(f'加载 {len(patents)} 条专利')
    return patents


def extract_keywords_from_name(patent_name: str) -> str:
    """从中文专利名提取英文关键词（用于 PubMed 检索）"""
    # 提取英文单词
    en_words = re.findall(r'[A-Za-z][A-Za-z0-9-]{2,}', patent_name)
    if en_words:
        return ' '.join(en_words[:5])

    # 中文专利名 → 提取关键生物医药术语的英文对应
    keyword_map = {
        '抗体': 'antibody', '疫苗': 'vaccine', '蛋白': 'protein',
        '基因': 'gene', '细胞': 'cell', '肿瘤': 'tumor',
        '癌': 'cancer', '病毒': 'virus', '免疫': 'immune',
        '药物': 'drug', '制剂': 'formulation', '纳米': 'nano',
        '多肽': 'peptide', '核酸': 'nucleic acid', '酶': 'enzyme',
        '受体': 'receptor', '抑制剂': 'inhibitor', '激酶': 'kinase',
        '干细胞': 'stem cell', '单克隆': 'monoclonal',
        '重组': 'recombinant', '转基因': 'transgenic',
        '发酵': 'fermentation', '提取': 'extraction',
        '检测': 'detection', '诊断': 'diagnosis',
        '治疗': 'therapy', '靶向': 'targeted',
        '冠状病毒': 'coronavirus', '新冠': 'COVID-19',
        '流感': 'influenza', '乙肝': 'hepatitis B',
        '糖尿病': 'diabetes', '阿尔茨海默': 'Alzheimer',
        '帕金森': 'Parkinson', '白血病': 'leukemia',
        '淋巴瘤': 'lymphoma', '乳腺癌': 'breast cancer',
        '肺癌': 'lung cancer', '肝癌': 'liver cancer',
        '胃癌': 'gastric cancer', '结直肠癌': 'colorectal cancer',
        '脂质体': 'liposome', '微球': 'microsphere',
        '生物标志物': 'biomarker', '分子标记': 'molecular marker',
        '克隆': 'clone', '表达': 'expression',
        '纯化': 'purification', '培养': 'culture',
        '灭活': 'inactivated', '减毒': 'attenuated',
    }

    keywords = []
    for cn, en in keyword_map.items():
        if cn in patent_name:
            keywords.append(en)
    return ' '.join(keywords[:5]) if keywords else ''


def search_pubmed(query: str, limit: int = 5) -> List[Dict]:
    """通过 PubMed E-utilities 检索论文"""
    try:
        # Step 1: esearch 获取 PMID 列表
        r = requests.get(PUBMED_SEARCH, params={
            'db': 'pubmed', 'term': query, 'retmax': limit, 'retmode': 'json'
        }, headers=HEADERS, timeout=15)

        if r.status_code != 200:
            return []

        ids = r.json().get('esearchresult', {}).get('idlist', [])
        if not ids:
            return []

        time.sleep(REQUEST_DELAY)

        # Step 2: esummary 获取摘要信息
        r2 = requests.get(PUBMED_SUMMARY, params={
            'db': 'pubmed', 'id': ','.join(ids), 'retmode': 'json'
        }, headers=HEADERS, timeout=15)

        if r2.status_code != 200:
            return []

        result = r2.json().get('result', {})
        papers = []

        for pid in ids:
            info = result.get(pid, {})
            if not isinstance(info, dict):
                continue

            title = info.get('title', '')
            authors_list = info.get('authors', [])
            authors = ';'.join(a.get('name', '') for a in authors_list[:10])
            year = info.get('pubdate', '')[:4]
            journal = info.get('fulljournalname', '') or info.get('source', '')
            doi = ''
            for aid in info.get('articleids', []):
                if aid.get('idtype') == 'doi':
                    doi = aid.get('value', '')
                    break

            if title:
                papers.append({
                    '论文名称': title,
                    '作者': authors,
                    '摘要': '',  # esummary 不返回摘要，需要 efetch
                    '发表年份': year,
                    '被引用次数': 0,  # PubMed 不直接提供引用次数
                    '期刊': journal,
                    'DOI': doi,
                    'PMID': pid,
                    '数据来源': 'PubMed',
                })

        return papers
    except Exception as e:
        logger.debug(f'PubMed error: {e}')
        return []


def build_cn_queries(patent: Dict) -> List[str]:
    """为一条专利构建 PubMed 检索查询"""
    queries = []
    inventors = patent['inventors']
    applicant = patent['applicant']
    patent_name = patent['patent_name']

    # 提取英文关键词
    keywords = extract_keywords_from_name(patent_name)

    # 策略1: 第一发明人 + 关键词（PubMed 能搜中文作者名）
    if inventors and keywords:
        queries.append(f'{inventors[0]}[Author] AND ({keywords})')

    # 策略2: 纯关键词搜索
    if keywords:
        queries.append(keywords)

    # 策略3: 专利名中的英文术语直接搜
    en_terms = re.findall(r'[A-Za-z][A-Za-z0-9-]{2,}', patent_name)
    if en_terms:
        queries.append(' '.join(en_terms[:6]))

    # 策略4: 申请人（可能是英文机构名）+ 关键词
    if applicant and keywords:
        # 检查申请人是否含英文
        en_in_app = re.findall(r'[A-Za-z]{3,}', applicant)
        if en_in_app:
            queries.append(f'{applicant} AND ({keywords})')

    return queries[:3]  # 最多3个查询


def paper_key(paper: Dict) -> str:
    title = paper.get('论文名称', '').strip().lower()
    return hashlib.md5(title.encode('utf-8')).hexdigest()


def fetch_papers(patents: List[Dict]) -> List[Dict]:
    all_papers = []
    seen: Set[str] = set()
    total = len(patents)
    skipped = 0

    for idx, patent in enumerate(patents):
        queries = build_cn_queries(patent)

        if not queries:
            skipped += 1
            if (idx + 1) % 100 == 0:
                logger.info(f'[{idx+1}/{total}] 跳过（无可用查询），累计跳过 {skipped}')
            continue

        logger.info(f'[{idx+1}/{total}] {patent["patid"]}')

        found = []
        for q in queries:
            if len(found) >= MAX_PAPERS_PER_PATENT:
                break

            results = search_pubmed(q, limit=5)
            time.sleep(REQUEST_DELAY)

            for p in results:
                k = paper_key(p)
                if k not in seen and p.get('论文名称'):
                    seen.add(k)
                    p['关联专利号'] = patent['patid']
                    p['关联发明人'] = ';'.join(patent['inventors'])
                    p['关联申请人'] = patent['applicant']
                    found.append(p)

        all_papers.extend(found[:MAX_PAPERS_PER_PATENT])

        if found:
            logger.info(f'  -> {len(found)} papers (total {len(all_papers)})')

        # 中间保存
        if (idx + 1) % 50 == 0:
            save_excel(all_papers, OUTPUT_XLSX)
            logger.info(f'  [checkpoint] {len(all_papers)} papers, skipped {skipped}')

    logger.info(f'跳过 {skipped} 条（无英文关键词可提取）')
    return all_papers


def save_excel(papers: List[Dict], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '中文相关论文'

    cols = ['论文名称', '作者', '摘要', '发表年份', '被引用次数',
            '期刊', 'DOI', 'PMID', '数据来源', '关联专利号', '关联发明人', '关联申请人']

    for ci, cn in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=cn)
        cell.font = Font(bold=True)

    for ri, paper in enumerate(papers, 2):
        for ci, cn in enumerate(cols, 1):
            ws.cell(row=ri, column=ci, value=paper.get(cn, ''))

    widths = [50, 30, 60, 10, 12, 30, 30, 12, 15, 20, 20, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description='中文相关论文检索（Phase 2 - PubMed）')
    parser.add_argument('--limit', type=int, default=0, help='处理前N条专利（0=全部）')
    args = parser.parse_args()

    logger.info('=== Phase 2: 中文相关论文检索（PubMed） ===')

    if not os.path.exists(INPUT_XLSX):
        logger.error(f'找不到: {INPUT_XLSX}')
        sys.exit(1)

    patents = load_patents(INPUT_XLSX)
    if args.limit > 0:
        patents = patents[:args.limit]

    logger.info(f'检索 {len(patents)} 条专利...')

    papers = fetch_papers(patents)
    save_excel(papers, OUTPUT_XLSX)

    logger.info(f'=== 完成: {len(papers)} 篇论文 -> {OUTPUT_XLSX} ===')


if __name__ == '__main__':
    main()
