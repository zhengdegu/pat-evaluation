#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从生物医药.xlsx 提取发明人+申请人，检索相关论文，结果存入 Excel。

检索策略：
1. 按发明人（前3位）+ 申请人单位组合检索
2. 使用百度学术网页检索（中文论文主力）+ CrossRef（英文补充）
3. 去重后存入 论文数据.xlsx

用法：
    python scripts/fetch_papers.py [--limit N]  # N=处理前N条专利，默认全部
"""

import os
import sys
import time
import json
import re
import logging
import hashlib
import argparse
from typing import List, Dict, Optional, Set
from urllib.parse import quote, urlencode

import requests
import openpyxl
from openpyxl.styles import Font

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE_DIR)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# === 配置 ===
INPUT_XLSX = os.path.join(BASE_DIR, '生物医药.xlsx')
OUTPUT_XLSX = os.path.join(BASE_DIR, 'data', '论文数据.xlsx')
MAX_PAPERS_PER_PATENT = 10
REQUEST_DELAY = 2.0  # 秒
MAX_INVENTORS_PER_PATENT = 3

HEADERS_BROWSER = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
}

HEADERS_API = {
    'User-Agent': 'PatEvaluation/1.0 (paper-fetcher; academic-research)',
    'Accept': 'application/json',
}

# CrossRef API
CROSSREF_API = 'https://api.crossref.org/works'

# Semantic Scholar API
S2_API = 'https://api.semanticscholar.org/graph/v1/paper/search'

# 百度学术搜索
BAIDU_XUESHU_URL = 'https://xueshu.baidu.com/s'


def load_patents(xlsx_path: str) -> List[Dict]:
    """从 Excel 加载专利数据"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else '' for h in rows[0]]

    patents = []
    for row in rows[1:]:
        d = dict(zip(headers, row))
        patid = str(d.get('专利号', '') or '').strip()
        patent_name = str(d.get('专利名', '') or '').strip()
        inv_str = str(d.get('发明人', '') or '').replace('；', ';')
        app_str = str(d.get('申请人', '') or '').replace('；', ';')

        inventors = [n.strip() for n in inv_str.split(';') if n.strip() and len(n.strip()) >= 2]
        applicant = app_str.split(';')[0].strip() if app_str else ''

        if not inventors:
            continue

        patents.append({
            'patid': patid,
            'patent_name': patent_name,
            'inventors': inventors[:MAX_INVENTORS_PER_PATENT],
            'applicant': applicant,
        })

    wb.close()
    logger.info(f'加载 {len(patents)} 条专利（共 {len(rows)-1} 条）')
    return patents


def search_baidu_xueshu(query: str, limit: int = 5) -> List[Dict]:
    """通过百度学术网页检索论文（解析 HTML）"""
    try:
        params = {'wd': query, 'pn': 0, 'ie': 'utf-8'}
        resp = requests.get(BAIDU_XUESHU_URL, params=params,
                   headers=HEADERS_BROWSER, timeout=15)
        if resp.status_code != 200:
            logger.debug(f'百度学术返回 {resp.status_code}')
            return []

        html = resp.text
        papers = []

        # 解析搜索结果（简单正则提取）
        # 百度学术结果在 <div class="result"> 块中
        result_blocks = re.findall(
            r'<div class="sc_content"[^>]*>(.*?)</div>\s*</div>\s*</div>',
            html, re.DOTALL
        )

        if not result_blocks:
            # 备用模式：提取标题和作者
            titles = re.findall(r'<a[^>]*class="sc_q"[^>]*>(.*?)</a>', html, re.DOTALL)
            if not titles:
                titles = re.findall(r'<h3[]*>(.*?)</a></h3>', html, re.DOTALL)

            for title_html in titles[:limit]:
                title = re.sub(r'<[^>]+>', '', title_html).strip()
                if title:
                    papers.append({
                        '论文名称': title,
                        '作者': '',
                        '摘要': '',
                        '发表年份': '',
                        '被引用次数': 0,
                        '期刊': '',
                        'DOI': '',
                        '数据来源': '百度学术',
                    })
            return papers

        for block in result_blocks[:limit]:
            # 提取标题
            title_match = re.search(r'<a[^>]*>(.*?)</a>', block)
            title = re.sub(r'<[^>]+>', '', title_match.group(1)).strip() if title_match else ''

            # 提取作者
            author_match = re.search(r'<span[^>]*class="sc_author"[^>]*>(.*?)</span>', block, re.DOTALL)
            authors = re.sub(r'<[^>]+>', '', author_match.group(1)).strip() if author_match else ''
            authors = authors.replace(' - ', ';').replace(',', ';').strip(' -')

            # 提取摘要
            abs_match = re.search(r'<span[^>]*class="sc_abstract"[^>]*>(.*?)</span>', block, re.DOTALL)
            abstract = re.sub(r'<[^>]+>', '', abs_match.group(1)).strip() if abs_match else ''

            # 提取年份
            year_match = re.search(r'(\d{4})', block)
            year = year_match.group(1) if year_match else ''

            # 提取引用次数
            cite_match = re.search(r'被引[：:]?\s*(\d+)', block)
            if not cite_match:
                cite_match = re.search(r'Cited by[：:]?\s*(\d+)', block)
            citations = int(cite_match.group(1)) if cite_match else 0

            # 提取期刊
            journal_match = re.search(r'<a[^>]*class="sc_q_journal"[^>]*>(.*?)</a>', block)
            journal = re.sub(r'<[^>]+>', '', journal_match.group(1)).strip() if journal_match else ''

            if title:
                papers.append({
                    '论文名称': title,
                    '作者': authors,
                    '摘要': abstract[:500],
                    '发表年份': year,
                    '被引用次数': citations,
                    '期刊': journal,
                    'DOI': '',
                    '数据来源': '百度学术',
                })

        return papers
    except Exception as e:
        logger.debug(f'百度学术检索异常: {e}')
        return []


def search_crossref(query: str, limit: int = 5) -> List[Dict]:
    """通过 CrossRef API 检索论文"""
    try:
        params = {
     uery': query,
            'rows': limit,
            'select': 'title,author,abstract,published-print,is-referenced-by-count,container-title,DOI',
        }
        resp = requests.get(CROSSREF_API, params=params, headers=HEADERS_API, timeout=15)
        if resp.status_code != 200:
            return []

        data = resp.json()
        papers = []
        for item in data.get('message', {}).get('items', []):
            title = item.get('title', [''])[0] if item.get('title') else ''
            authors = ';'.join(
                f"{a.get('family', '')} {a.get('given', '')}".strip()
                for a in item.get('author', [])
            )
            abstract_raw = item.get('abstract', '')
            abstract = re.sub(r'<[^>]+>', '', abstract_raw)[:500] if abstract_raw else ''
            pub_date = item.get('published-print', {}).get('date-parts', [[None]])[0]
            year = pub_date[0] if pub_date and pub_date[0] else ''
            journal = item.get('container-title', [''])[0] if item.get('container-title') else ''

            papers.append({
                '论文名称': title,
                '作者': authors,
                '摘要': abstract,
                '发表年份': y              '被引用次数': item.get('is-referenced-by-count', 0),
                '期刊': journal,
                'DOI': item.get('DOI', ''),
                '数据来源': 'CrossRef',
            })
        return papers
    except Exception as e:
        logger.debug(f'CrossRef 检索异常: {e}')
        return []


def search_semantic_scholar(query: str, limit: int = 5) -> List[Dict]:
    """通过 Semantic Scholar API 检索论文（有限流，作为补充）"""
    try:
        params = {
            'query': query,
            'limit': limit,
            'fields': 'title,authors,abstract,year,citationCount,journal,externalIds',
        }
        resp = requests.get(S2_API, params=params, headers=HEADERS_API, timeout=15)
        if resp.status_code == 429:
            logger.debug('S2 限流，跳过')
            return []
        if resp.status_code != 200:
            return []

        data = resp.json()
        papers = []
        for item in data.get('data', []):
            authors = ';'.join(a.get('name', '') for a in item.get('authors', []))
            doi = (item.get('externalIds') or {}).get('DOI', '')
            journal = (item.get('journal') or {}).get('name', '')
            papers.append({
                '论文名称': item.get('title',n                '作者': author                '摘要': (item.get('abstract') or '')[:500],
                '发表年份': item.get('year', ''),
                '被引用次数': item.get('citationCount', 0),
                '期刊': journal,
                'DOI': doi,
                '数据来源': 'Semantic Scholar',
            })
        return papers
    except Exception as e:
        logger.debug(f'S2 检索异常: {e}')
        return []


def build_queries(patent: Dict) -> List[str]:
    """为一条专利构建检索查询"""
    queries = []
    applicant = patent['applicant']
    inventors = patent['inventors']

    # 策略1: 第一发明人 + 申请人单位
    if inventors and applicant:      queries.append(f'{inventors[0]} {applicant}')

    # 策略2: 前两个发明人组合
    if len(inventors) >= 2:
        queries.append(f'{inventors[0]} {inventors[1]}')

    # 策略3: 第一发明人 + 专利名关键词
    if inventors:
        patent_kw = patent['patent_name'][:20]
        queries.append(f'{inventors[0]} {patent_kw}')

    return queries


def paper_key(paper: Dict) -> str:
    """论文去重 key"""
    title = paper.get('论文名称', '').strip().lower()
    return hashlib.md5(title.encode('utf-8')).hexdigest()


def fetch_papers_for_patenents: List[Dict]) -> List[Dict]:
    """批量检索论文"""
    all_papers = []
    seen_keys: Set[str] = set()
    total = len(patents)
    s2_available = True  # S2 限流后关闭

    for idx, patent in enumerate(patents):
        logger.info(f'[{idx+1}/{total}] {patent["patid"]} - {patent["inventors"][0]}')

        queries = build_queries(patent)
        patent_papers = []

        for q in queries:
            if len(patent_papers) >= MAX_PAPERS_PER_PATENT:
                break

            # 主力：百度学术（中文论文）
            results = search_baidu_xueshu(q, limit=5)
            time.sleep(REQUEST_DELAY)

            # 补充：CrossRef（英文论文）
            if lensults) < 3:
                results.extend(search_crossref(q, limit=5))
                time.sleep(REQUEST_DELAY)

            # 再补充：Semantic Scholar（如果可用）
            if len(results) < 3 and s2_available:
                s2_results = search_semantic_scholar(q, limit=3)
                if s2_results:
                    results.extend(s2_results)
                else:
                    s2_available = False  # 可能被限流了
                time.sleep(REQUEST_DELAY)

            for paper in results:
                key = paper_key(paper)
                if key not in seen_keys and paper.get('论文名称'):
              keys.add(key)
                    paper['关联专利号'] = patent['patid']
                    paper['关联发明人'] = ';'.join(patent['inventors'])
                    paper['关联申请人'] = patent['applicant']
                    patent_papers.append(paper)

        all_papers.extend(patent_papers[:MAX_PAPERS_PER_PATENT])

        if patent_papers:
            logger.info(f'  -> {len(patent_papers)} 篇（累计 {len(all_papers)}）')
        else:
            logger.info(f'  -> 未找到论文')

        # 每50条保存一次中间结果
        if (idx + 1) % 50 == 0:
            save_to_excel(all_papers, OUTPUT_XLSX)
            logger.info(f'  [中间保存] {len(all_papers)} 篇')

    return all_papers


def save_to_excel(papers: List[Dict], output_path: str):
    """保存论文数据到 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '论文数据'

    columns = [
        '论文名称', '作者', '摘要', '发表年份', '被引用次数',
        '期刊', 'DOI', '数据来源', '关联专利号', '关联发明人', '关联申请人'
    ]

    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font

    for row_idx, paper in enumerate(papers, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = paper.get(col_name, '')
            ws.cell(row=row_idx, column=col_idx, value=value)

    col_widths = [50, 30, 60, 10, 12, 30, 30, 15, 20, 20, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description='论文数据检索工具')
    parser.add_argument('--limit', type=int, default=0, help='只处理前N条专利（0=全部）')
    args = parser.parse_args()

    logger.info('=== 论文数据检索工具 ===')

    if not os.path.exists(INPU
        logger.error(f'找不到输入文件: {INPUT_XLSX}')
        sys.exit(1)

    patents = load_patents(INPUT_XLSX)
    if args.limit > 0:
        patents = patents[:args.limit]
        logger.info(f'限制处理前 {args.limit} 条')

    logger.info(f'开始检索 {len(patents)} 条专利的相关论文...')
    est_minutes = len(patents) * 3 * REQUEST_DELAY / 60
    logger.info(f'预计耗时: {est_minutes:.0f} 分钟')

    papers = fetch_papers_for_patents(patents)
    save_to_excel(papers, OUTPUT_XLSX)

    logger.info(f'=== 完成 ===')
    logger.info(f'共检索到 {len(papers)} 篇去重论文')
    logger.info(f'输出文件: {OUTPUT_XLSX}')


if __name__ == '__main__':
    main()
