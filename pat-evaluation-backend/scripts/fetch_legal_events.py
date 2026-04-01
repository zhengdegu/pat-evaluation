#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Fetch legal events from Google Patents for all patents in the Excel file.
Extract transfer/license/pledge trade records.

Usage:
    python scripts/fetch_legal_events.py --test CN1197876C
    python scripts/fetch_legal_events.py --limit 5
    python scripts/fetch_legal_events.py --resume
"""

import os
import sys
import json
import time
import random
import re
import argparse

import requests
from bs4 import BeautifulSoup
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "..", "生物医药.xlsx")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "output")
EVENTS_FILE = os.path.join(OUTPUT_DIR, "legal_events.json")
TRADES_FILE = os.path.join(OUTPUT_DIR, "patent_trades.json")
PROGRESS_FILE = os.path.join(OUTPUT_DIR, "fetch_progress.json")

GP_URL = "https://patents.google.com/patent/{}/en"

TRANSFER_CODES = {"TR01", "C41", "ASS"}
LICENSE_CODES = {"EE01", "EM01"}
PLEDGE_CODES = {"PM01", "PE01", "PC01", "PLDC"}
TRADE_CODES = TRANSFER_CODES | LICENSE_CODES | PLEDGE_CODES

sess = requests.Session()
sess.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "en-US,en;q=0.9",
})


def jitter(lo=1.5, hi=3.5):
    time.sleep(random.uniform(lo, hi))


def load_patents(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    result = []
    for r in rows[1:]:
        d = dict(zip(hdr, r))
        pn = str(d.get(hdr[0], "") or "").strip()
        pub = str(d.get(hdr[8], "") or "").strip()
        app = str(d.get(hdr[4], "") or "").strip()
        own = str(d.get(hdr[5], "") or "").strip()
        tv = 0.0
        try:
            tv = float(d.get(hdr[21], 0) or 0)
        except Exception:
            pass
        if pn and pub:
            result.append(dict(pn=pn, pub=pub, app=app, own=own, tv=tv))
    wb.close()
    return result


def fetch_events(pub_no):
    pub_no = re.sub(r"\s+", "", pub_no.strip())
    url = GP_URL.format(pub_no)
    try:
        resp = sess.get(url, timeout=20)
        if resp.status_code == 404:
            alt = re.sub(r"[A-Za-z]$", "", pub_no)
            if alt != pub_no:
                resp = sess.get(GP_URL.format(alt), timeout=20)
        if resp.status_code != 200:
            return None
    except requests.RequestException as e:
        print("  [ERR] {}: {}".format(pub_no, e))
        return None

    soup = BeautifulSoup(resp.text, "lxml")
    events = []

    for row in soup.select('tr[itemprop="legalEvents"]'):
        ev = _parse_row(row)
        if ev:
            events.append(ev)
    if events:
        return events

    for sec in soup.find_all("section", itemprop="legalEvents"):
        for row in sec.find_all("tr"):
            ev = _parse_row(row)
            if ev:
                events.append(ev)
    return events if events else None


def _parse_row(row):
    cells = row.find_all(["td", "th"])
    if len(cells) < 3:
        return None
    dt = cells[0].get_text(strip=True)
    cd = cells[1].get_text(strip=True)
    tt = cells[2].get_text(strip=True) if len(cells) > 2 else ""
    ds = cells[3].get_text(" | ", strip=True) if len(cells) > 3 else ""
    if not dt or not cd:
        return None
    return {"date": dt, "code": cd, "title": tt, "description": ds}


def parse_details(desc):
    d = {}
    if not desc:
        return d
    # Google Patents uses " | : | " as separator in description text
    # Support both "Key: Value" and "Key | : | Value" formats
    m = re.search(r"Effective date[^:]*(?::\s*|\s*\|\s*:\s*\|\s*)(\d{8})", desc)
    if m:
        s = m.group(1)
        d["effective_date"] = "{}-{}-{}".format(s[:4], s[4:6], s[6:8])
    m = re.search(r"Patentee after\s*\|\s*:\s*\|\s*(.+?)(?:\s*\|\s*Address|\s*$)", desc)
    if not m:
        m = re.search(r"Patentee after:\s*(.+?)(?:\||Address|$)", desc)
    if m:
        d["new_owner"] = m.group(1).strip().rstrip("|").strip()
    m = re.search(r"Patentee before\s*\|\s*:\s*\|\s*(.+?)(?:\s*\|\s*|\s*$)", desc)
    if not m:
        m = re.search(r"Patentee before:\s*(.+?)(?:\||Address|$)", desc)
    if m:
        d["original_owner"] = m.group(1).strip().rstrip("|").strip()
    m = re.search(r"FORMER OWNER:\s*(.+?)(?:\||Effective|$)", desc)
    if m and "original_owner" not in d:
        d["original_owner"] = m.group(1).strip()
    m = re.search(r"Owner name\s*\|\s*:\s*\|\s*(.+?)(?:\s*\|\s*Free|\s*\|\s*$|\s*$)", desc)
    if not m:
        m = re.search(r"Owner name:\s*(.+?)(?:\||Free|$)", desc)
    if m and "new_owner" not in d:
        d["new_owner"] = m.group(1).strip().rstrip("|").strip()
    m = re.search(r"Assignee\s*\|\s*:\s*\|\s*(.+?)(?:\s*\|\s*Assignor|\s*\|\s*$|\s*$)", desc)
    if not m:
        m = re.search(r"Assignee:\s*(.+?)(?:\||Assignor|$)", desc)
    if m:
        d["licensee"] = m.group(1).strip().rstrip("|").strip()
    m = re.search(r"Assignor\s*\|\s*:\s*\|\s*(.+?)(?:\s*\|\s*Contract|\s*\|\s*$|\s*$)", desc)
    if not m:
        m = re.search(r"Assignor:\s*(.+?)(?:\||Contract|$)", desc)
    if m:
        d["licensor"] = m.group(1).strip().rstrip("|").strip()
    m = re.search(r"License type\s*\|\s*:\s*\|\s*(.+?)(?:\s*\|\s*|\s*$)", desc)
    if not m:
        m = re.search(r"License type:\s*(.+?)(?:\||$)", desc)
    if m:
        d["license_type"] = m.group(1).strip().rstrip("|").strip()
    m = re.search(r"Pledgee\s*\|\s*:\s*\|\s*(.+?)(?:\s*\|\s*Pledgor|\s*\|\s*$|\s*$)", desc)
    if not m:
        m = re.search(r"Pledgee:\s*(.+?)(?:\||Pledgor|$)", desc)
    if m:
        d["pledgee"] = m.group(1).strip().rstrip("|").strip()
    m = re.search(r"Pledgor\s*\|\s*:\s*\|\s*(.+?)(?:\s*\|\s*Reg|\s*\|\s*$|\s*$)", desc)
    if not m:
        m = re.search(r"Pledgor:\s*(.+?)(?:\||Reg)", desc)
    if m:
        d["pledgor"] = m.group(1).strip().rstrip("|").strip()
    return d


def extract_trades(pat_no, pub_no, events):
    trades = []
    for ev in events:
        code = ev.get("code", "")
        if code not in TRADE_CODES:
            continue
        if code in TRANSFER_CODES:
            tt = "转让"
        elif code in LICENSE_CODES:
            tt = "许可"
        elif code in PLEDGE_CODES:
            tt = "质押"
        else:
            continue
        det = parse_details(ev.get("description", ""))
        orig = det.get("original_owner", det.get("licensor", det.get("pledgor", "")))
        new = det.get("new_owner", det.get("licensee", det.get("pledgee", "")))
        trades.append({
            "专利号": pat_no,
            "公开号": pub_no,
            "交易类型": tt,
            "交易日期": det.get("effective_date", ev.get("date", "")),
            "交易金额": None,
            "原权利人": orig,
            "新权利人": new,
            "许可类型": det.get("license_type", ""),
            "数据来源": "Google Patents",
            "法律事件代码": code,
            "法律事件标题": ev.get("title", ""),
            "原始描述": ev.get("description", ""),
        })
    return trades


def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_json(path, default=None):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return default if default is not None else {}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--resume", action="store_true")
    ap.add_argument("--test", type=str, default="")
    args = ap.parse_args()

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if args.test:
        print("Testing: " + args.test)
        evts = fetch_events(args.test)
        if evts:
            print("Found {} legal events:".format(len(evts)))
            for e in evts:
                mark = " <<< TRADE" if e["code"] in TRADE_CODES else ""
                print("  {} | {} | {}{}".format(e["date"], e["code"], e["title"], mark))
                if e["code"] in TRADE_CODES:
                    print("    " + e["description"][:300])
            tds = extract_trades("TEST", args.test, evts)
            if tds:
                print("\nExtracted {} trades:".format(len(tds)))
                for t in tds:
                    print(json.dumps(t, ensure_ascii=False, indent=2))
        else:
            print("No events found")
        return

    print("Loading " + EXCEL_PATH)
    patents = load_patents(EXCEL_PATH)
    print("Loaded {} patents".format(len(patents)))

    if args.resume:
        prog = load_json(PROGRESS_FILE, {"completed": [], "failed": [], "last_index": 0})
    else:
        prog = {"completed": [], "failed": [], "last_index": 0}
    done = set(prog["completed"])

    all_ev = load_json(EVENTS_FILE, {}) if args.resume else {}
    all_tr = load_json(TRADES_FILE, []) if args.resume else []

    total = min(len(patents), args.limit) if args.limit > 0 else len(patents)
    ok_n = 0
    fail_n = 0

    for i, pat in enumerate(patents[:total]):
        pn = pat["pn"]
        pub = pat["pub"]
        if pn in done:
            continue
        sys.stdout.write("[{}/{}] {} ({})... ".format(i + 1, total, pn, pub))
        sys.stdout.flush()

        evts = fetch_events(pub)
        if evts is not None:
            all_ev[pn] = {
                "专利号": pn,
                "公开号": pub,
                "申请人": pat["app"],
                "当前权利人": pat["own"],
                "events": evts,
            }
            tds = extract_trades(pn, pub, evts)
            if tds:
                all_tr.extend(tds)
                print("OK ({} events, {} trades)".format(len(evts), len(tds)))
            else:
                print("OK ({} events)".format(len(evts)))
            ok_n += 1
            prog["completed"].append(pn)
            done.add(pn)
        else:
            print("FAIL")
            fail_n += 1
            prog["failed"].append(pn)

        prog["last_index"] = i
        if (i + 1) % 10 == 0:
            save_json(EVENTS_FILE, all_ev)
            save_json(TRADES_FILE, all_tr)
            save_json(PROGRESS_FILE, prog)
            print("  --- {} ok, {} fail, {} trades ---".format(ok_n, fail_n, len(all_tr)))

        jitter(1.0, 2.5)

    save_json(EVENTS_FILE, all_ev)
    save_json(TRADES_FILE, all_tr)
    save_json(PROGRESS_FILE, prog)

    total_ev = sum(len(v["events"]) for v in all_ev.values())
    print("\nDone! {} ok, {} fail, {} events, {} trades".format(ok_n, fail_n, total_ev, len(all_tr)))
    print("Output: " + EVENTS_FILE)
    print("Trades: " + TRADES_FILE)


if __name__ == "__main__":
    main()
